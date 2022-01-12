# -*- coding: utf-8 -*-
from telnetlib import EC

import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook
from bs4 import BeautifulSoup
from urllib.request import urlopen
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


driver = webdriver.Chrome(executable_path='chromedriver')
driver.maximize_window()

write_wb = Workbook();
write_ws = write_wb.active

count = 1

url = 'https://aihub.or.kr/aihub-data/healthcare/about'
driver.get(url)
write_ws.cell(1, 1,'데이터셋 명')
write_ws.cell(1, 2,'데이터 소개')
write_ws.cell(1, 3,'데이터 유형')
write_ws.cell(1, 4,'데이터 제공 기관')
write_ws.cell(1, 5,'데이터 량')
write_ws.cell(1, 6,'데이터 구축 년도')
write_ws.cell(1, 7,'데이터 수정 일자')
write_ws.cell(1, 8,'데이터 키워드')
write_ws.cell(1, 9,'저작권 및 이용정책')
write_ws.cell(1,10,'URL')
write_ws.cell(1,11,'데이터 카테고리')
# //*[@id="block-views-block-8dae-bunlyu-block-7"]/div/div/div/div/ul/li[1]/div/span/a
for i in range (35) :
    xpathhead = '//*[@id="block-views-block-8dae-bunlyu-block-7"]/div/div/div/div/ul/li['
    xpathmiddle = str(i+1)
    xpathtail = ']/div/span/a'
    # //*[@id="block-views-block-8dae-bunlyu-block-5"]/div/div/div/div/ul/li[1]/div/span/a
    xpath = xpathhead + xpathmiddle + xpathtail
    hrText = driver.find_element(By.XPATH, xpath).get_attribute('href')
    print(hrText)
    driver.get(hrText)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    dataName = driver.find_element(By.XPATH,
                                   '//*[@id="container"]/article/div[3]/div/div/div/div[1]/span/div/div/table/tbody/tr[1]/td').text
    # 데이터셋 명
    dataInfo = driver.find_element(By.XPATH,
                                   '//*[@id="container"]/article/div[3]/div/div/div/div[1]/span/div/div/table/tbody/tr[8]/td').text
    # 데이터 소개
    dataType = driver.find_element(By.XPATH,
                                   '//*[@id="container"]/article/div[3]/div/div/div/div[1]/span/div/div/table/tbody/tr[2]/td[2]').text
    # 데이터 유형
    dataFrom = driver.find_element(By.XPATH,
                                   '//*[@id="container"]/article/div[3]/div/div/div/div[1]/span/div/div/table/tbody/tr[3]/td[1]').text
    # 데이터 제공 기관
    dataCount = driver.find_element(By.XPATH,
                                    '//*[@id="container"]/article/div[3]/div/div/div/div[1]/span/div/div/table/tbody/tr[6]/td[1]').text
    # 데이터 량
    dataBirth = driver.find_element(By.XPATH,
                                    '//*[@id="container"]/article/div[3]/div/div/div/div[1]/span/div/div/table/tbody/tr[6]/td[2]').text
    # 데이터 구축 년도
    dataUpdateDate = driver.find_element(By.XPATH,
                                         '//*[@id="container"]/article/div[3]/div/div/div/div[1]/span/div/div/table/tbody/tr[7]/td[2]').text
    # 데이터 수정 일자
    dataKeyWord = driver.find_element(By.XPATH,
                                      '//*[@id="container"]/article/div[3]/div/div/div/div[1]/span/div/div/table/tbody/tr[9]/td').text
    # 데이터 키워드
    dataCopyright = driver.find_element(By.XPATH,
                                        '//*[@id="container"]/article/div[3]/div/div/div/div[1]/span/div/div/table/tbody/tr[10]/td').text
    # 저작권 및 이용정책
    count += 1
    write_ws.cell(count, 1, dataName)
    write_ws.cell(count, 2, dataInfo)
    write_ws.cell(count, 3, dataType)
    write_ws.cell(count, 4, dataFrom)
    write_ws.cell(count, 5, dataCount)
    write_ws.cell(count, 6, dataBirth)
    write_ws.cell(count, 7, dataUpdateDate)
    write_ws.cell(count, 8, dataKeyWord)
    write_ws.cell(count, 9, dataCopyright)
    thisUrl = driver.current_url
    write_ws.cell(count, 10, thisUrl)
    write_ws.cell(count,11,'헬스케어')
    driver.back()


write_wb.save('AI_health.xlsx')

