
#coinmarketcap
import time
from selenium import webdriver
from selenium.webdriver.common.by import By

from selenium.webdriver import ActionChains
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook

def getPrice()  -> int:
    url = "https://coinmarketcap.com/ko/currencies/solana/"
    driver.get(url)

    price_sol = driver.find_element(By.XPATH,
                                    '//*[@id="__next"]/div/div[1]/div[2]/div/div[1]/div[2]/div/div[2]/div[1]/div/span').text
    price_sol = price_sol[1:len(price_sol) - 3]
    temp = ''
    for i in price_sol:
        if i != ',':
            temp += i

    price_sol = temp
    print(price_sol)

    url = "https://coinmarketcap.com/ko/currencies/binance-usd/"
    driver.get(url)

    price_busd = driver.find_element(By.XPATH,
                                     '//*[@id="__next"]/div/div[1]/div[2]/div/div[1]/div[2]/div/div[2]/div[1]/div/span').text
    price_busd = price_busd[1:len(price_busd) - 3]
    temp = ''
    for i in price_busd:
        if i != ',':
            temp += i
    price_busd = temp
    print(price_busd)
    return price_sol,price_busd

def toWon(s: str,sol :int,bus : int)  -> int:
    cointype = s[len(s)-3:]
    print(cointype)
    temp = ''
    for i in s:
        if 48<=ord(i)<=57:
            temp += i
        elif ord(i)==46:
            break
    print("value :"+temp)
    if cointype =='SOL':
        temp = int(temp)
        value = temp*sol
        return value
    else:
        temp = int(temp)
        value = temp*bus
        return value


driver = webdriver.Chrome(executable_path='chromedriver')
driver.maximize_window()



sol,busd =getPrice()
sol = int(sol)
busd = int(busd)
print(sol)
print(busd)

url = "https://coinmarketcap.com/ko/nft/collections/"
driver.get(url)


page_height = driver.execute_script("return document.body.scrollHeight")

driver.find_element(By.XPATH ,'//*[@id="__next"]/div/div[1]/div[2]/div/div'
                              '/div[1]/div/button[4]/span').click()

interval =1

action = ActionChains(driver)

write_wb = Workbook();
write_ws = write_wb.active
count = 0

pcount =0;

for page in range(0,9):
    url = 'https://coinmarketcap.com/ko/nft/collections/?page=' + str(page+1)
    driver.get(url)
    driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[1]/div[2]/div/div'
                                  '/div[1]/div/button[4]/span').click()
    time.sleep(2)
    for i in range(0,100):
        count+=1
        head = '//*[@id="__next"]/div/div[1]/div[2]/div/div/div[3]/table/tbody/tr['
        title_tail = ']/td[2]'
        title_xpath = head + str(i+1) + title_tail
        print(count)
        # //*[@id="__next"]/div/div[1]/div[2]/div/div/div[3]/table/tbody/tr[1]/td[5] floor price
        # //*[@id="__next"]/div/div[1]/div[2]/div/div/div[3]/table/tbody/tr[1]/td[7] sales
        # //*[@id="__next"]/div/div[1]/div[2]/div/div/div[3]/table/tbody/tr[1]/td[8] properties
        # //*[@id="__next"]/div/div[1]/div[2]/div/div/div[3]/table/tbody/tr[2]/td[9] owner

        volume_tail = ']/td[3]'
        volume_xpath = head + str(i+1) + volume_tail

        floorPrice_tail = ']/td[5]'
        floorPrice_xpath = head + str(i+1) + floorPrice_tail

        sales_tail = ']/td[7]'
        sales_xpath = head + str(i+1) + sales_tail

        properties_tail = ']/td[8]'
        properties_xpath = head +str(i+1) +properties_tail

        owner_tail = ']/td[9]'
        owner_xpath = head + str(i+1) + owner_tail

        try :
            title = driver.find_element(By.XPATH, title_xpath)
            volume = driver.find_element(By.XPATH,volume_xpath)
            floorPrice = driver.find_element(By.XPATH, floorPrice_xpath)
            sales = driver.find_element(By.XPATH, sales_xpath)
            properties = driver.find_element(By.XPATH, properties_xpath)
            owner = driver.find_element(By.XPATH, owner_xpath)
        except NoSuchElementException :
            continue
        volume = toWon(volume.text,sol,busd)
        if(floorPrice.text!='--'):
            floorPrice = toWon(floorPrice.text,sol,busd)
            write_ws.cell(count,3,str(floorPrice))
        else:
            write_ws.cell(count, 3, str(0))
        action.move_to_element(title).perform()
        write_ws.cell(count, 1, title.text)
        write_ws.cell(count, 2, str(volume))
        write_ws.cell(count, 4, sales.text)
        write_ws.cell(count, 5, properties.text)
        write_ws.cell(count, 6, owner.text)
    while True:
        driver.execute_script(
            "window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(interval)

        curr_height = driver.execute_script("return document.body.scrollHeight")
        if curr_height == page_height:
            break
        page_height = curr_height
    time.sleep(5)

write_wb.save('900.xlsx')
