# -*- coding: utf-8 -*-

import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from bs4 import BeautifulSoup

driver = webdriver.Chrome(executable_path='chromedriver')
driver.maximize_window()
write_wb = Workbook();
write_ws = write_wb.active

write_ws.cell(1, 1, '데이터셋 명')
write_ws.cell(1, 2, '데이터 소개')
write_ws.cell(1, 3, '공개일자')
write_ws.cell(1, 4, '최신수정일자')
write_ws.cell(1, 5, '갱신주기')
write_ws.cell(1, 6, '분류')
write_ws.cell(1, 7, '원본시스템')
write_ws.cell(1, 8, '저작권자')
write_ws.cell(1, 9, '제공기관')
write_ws.cell(1, 10, '제공부서')
write_ws.cell(1, 11, '담당자')
write_ws.cell(1, 12, '원본형태')
write_ws.cell(1, 13, '제3저작권자')
write_ws.cell(1, 14, '라이센스')
write_ws.cell(1, 15, '관련 태그')
write_ws.cell(1,16,'URL')
count = 1

url = 'https://data.seoul.go.kr/dataList/datasetList.do'
driver.get(url)

for c in range(15):
    driver.find_element(By.XPATH, '//*[@id="datasetVO"]/div[2]/div/section/div[2]/div/div/button[13]').click()
for k in range(5):
    for j in range(10):
        for i in range(10):
            count += 1
            # //*[@id="datasetVO"]/div[2]/div/section/div[2]/dl[1]/dt/a
            # //*[@id="datasetVO"]/div[2]/div/section/div[2]/dl[3]/dt/a
            xpathhead = '//*[@id="datasetVO"]/div[2]/div/section/div[2]/dl['
            xpathmiddle = str(i + 1)
            xpathtail = ']/dt/a'
            xpath = xpathhead + xpathmiddle + xpathtail
            driver.find_element(By.XPATH,xpath).click()
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            print(count)
            datasetname = driver.find_element(By.XPATH,'//*[@id="frm"]/div[1]/h1').text
            # 데이터셋 명
            datainfo = driver.find_element(By.XPATH,'//*[@id="frm"]/div[1]/div[1]/p').text
            # 데이터 소개
            opendate = driver.find_element(By.XPATH,'//*[@id="frm"]/div[3]/div[2]/table/tbody/tr[1]/td[1]/span').text
            # 공개일자
            lastupdate = driver.find_element(By.XPATH,'//*[@id="frm"]/div[3]/div[2]/table/tbody/tr[1]/td[2]/span').text
            # 최신수정일자
            updatecycle = driver.find_element(By.XPATH,'//*[@id="frm"]/div[3]/div[2]/table/tbody/tr[2]/td[1]').text
            # 갱신주기
            category = driver.find_element(By.XPATH,'//*[@id="frm"]/div[3]/div[2]/table/tbody/tr[2]/td[2]').text
            # 분류
            originsystem = driver.find_element(By.XPATH,'//*[@id="frm"]/div[3]/div[2]/table/tbody/tr[3]/td[1]').text
            # 원본시스템
            copyowner = driver.find_element(By.XPATH,'//*[@id="frm"]/div[3]/div[2]/table/tbody/tr[3]/td[2]').text
            # 저작권자
            producer = driver.find_element(By.XPATH,'//*[@id="frm"]/div[3]/div[2]/table/tbody/tr[4]/td[1]').text
            # 제공기관
            producerpart = driver.find_element(By.XPATH,'//*[@id="frm"]/div[3]/div[2]/table/tbody/tr[4]/td[2]').text
            # 제공부서
            manager = driver.find_element(By.XPATH, '//*[@id="frm"]/div[3]/div[2]/table/tbody/tr[5]/td').text
            # 담당자
            originaldata = driver.find_element(By.XPATH, '//*[@id="frm"]/div[3]/div[3]/table/tbody/tr[9]/td').text
            # 원본형태
            copyowner3 = driver.find_element(By.XPATH, '//*[@id="frm"]/div[3]/div[2]/table/tbody/tr[6]/td[2]').text
            # 제3저작권자
            license = driver.find_element(By.XPATH, '//*[@id="frm"]/div[3]/div[2]/table/tbody/tr[7]/td/div').text
            # 라이센스
            tag = driver.find_element(By.XPATH, '//*[@id="frm"]/div[3]/div[2]/table/tbody/tr[8]/td').text
            # 관련태그

            thisUrl = driver.current_url
            write_ws.cell(count, 1, datasetname)
            write_ws.cell(count, 2, datainfo)
            write_ws.cell(count, 3, opendate)
            write_ws.cell(count, 4, lastupdate)
            write_ws.cell(count, 5, updatecycle)
            write_ws.cell(count, 6, category)
            write_ws.cell(count, 7, originsystem)
            write_ws.cell(count, 8, copyowner)
            write_ws.cell(count, 9, producer)
            write_ws.cell(count, 10, producerpart)
            write_ws.cell(count, 11, manager)
            write_ws.cell(count, 12, originaldata)
            write_ws.cell(count, 13, copyowner3)
            write_ws.cell(count, 14, license)
            write_ws.cell(count, 15, tag)
            write_ws.cell(count, 16, thisUrl)
            driver.back()

        xhead = '//*[@id="datasetVO"]/div[2]/div/section/div[2]/div/div/button['
        xmiddle = str(j + 4)
        xtail = ']'
        next_page_xpath = xhead + xmiddle + xtail
        driver.find_element(By.XPATH, next_page_xpath).click()
write_wb.save('seoul_1500to2000.xlsx')



# 다음 페이지 버튼 클릭
# 서울 열린데이터 광장 url 따기